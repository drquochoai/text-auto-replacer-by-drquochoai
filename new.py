import atexit
import logging
import configparser
import sys
import os
import threading
import openpyxl
import requests
import keyboard
import time
from pynput import mouse
import pyperclip
import win32clipboard
import tkinter as tk
from tkinter import messagebox, ttk
import webbrowser
from packaging import version
from ttkthemes import ThemedStyle

class TextReplacerApp:
    def __init__(self):
        # --- StarterConfig.py content ---
        # Set up logging
        logging.basicConfig(level=logging.INFO)

        # Constants
        self.VERSION = "0.0.3"
        self.TIME_INTERVAL_CLIPBOARD_CHECK = 0.8
        self.TIME_INTERVAL_KEYBOARD_CHECK = 0.5
        self.BACKUP_XLSX_PATH = "backup_replacement_data.xlsx"
        self.LANGUAGES_FOLDER = "languages"
        # --- client_config.py content ---
        self.HTTP_TIMEOUT = 30
        self.UPDATE_URLS = ['http://localhost:8000/updates']
        self.DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQKz1tiROR-S8zLK6YkrR5OPsvsuJAEVi1uC1ecTKk5-MLC-g6_jzIvSwAUNdnN5kyuIzbvU2DkmH1g/pub?output=xlsx"
        self.LINK_EDIT_FILE = "https://docs.google.com/spreadsheets/d/16uVFfVMKR7jVXA70g4BCo8KAE7iZVYnJT48oTpD1Z-4/edit?gid=0#gid=0"
        self.DEFAULT_BEFORE_REPLACEMENT = ""
        self.DEFAULT_AFTER_REPLACEMENT = " "
        self.LINK_LANGUAGE_FILE = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSVoAsKwGTxQyR16vv8rLTwEx07N4OxZpK7qDql-tnb3sc3sOe6YCsJ549C3xFMNfMLO6Knn2I5By_Q/pub?output=xlsx"

        # Global variables
        self.last_mouse_position = (0, 0)
        self.mouse_moved_significantly = False
        self.is_paused = False
        self.keyboard_thread_running = False
        self.current_replacement_data = {}
        self.stop_event = threading.Event()  # Global event to signal thread stop

        # Load settings from .ini file
        self.config = configparser.ConfigParser()
        self.config.read('settings.ini')

        # Load language settings
        self.language_config = configparser.ConfigParser()

        # Get settings or use defaults
        self.SHEET_URL = self.config.get('Settings', 'sheet_url', fallback=self.DEFAULT_SHEET_URL)
        self.BEFORE_REPLACEMENT = self.config.get('Settings', 'before_replacement', fallback=self.DEFAULT_BEFORE_REPLACEMENT)
        self.AFTER_REPLACEMENT = self.config.get('Settings', 'after_replacement', fallback=self.DEFAULT_AFTER_REPLACEMENT)

        # Track previous values of the widgets
        self.previous_sheet_url = self.SHEET_URL
        self.previous_before_replacement = self.BEFORE_REPLACEMENT
        self.previous_after_replacement = self.AFTER_REPLACEMENT


        # --- tkGUI.py content ---
        # Create the GUI
        self.root = tk.Tk()
        self.style = ThemedStyle(self.root)
        self.style.set_theme('radiance')

        self.root.title("Text Replacer by drquochoai")
        self.style.configure("TButton", width=20)
        self.style.configure("TOptionMenu", width=15)

        # Custom font
        custom_font = ("Roboto", 10)

        # Labels
        self.sheet_url_label = ttk.Label(self.root, text="Sheet URL:", font=custom_font)
        self.sheet_url_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        self.before_replacement_label = ttk.Label(self.root, text="Before Replacement:", font=custom_font)
        self.before_replacement_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")

        self.after_replacement_label = ttk.Label(self.root, text="After Replacement:", font=custom_font)
        self.after_replacement_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")

        # Entry fields
        self.sheet_url_text = tk.Text(self.root, width=50, height=3, font=custom_font, relief="solid", bd=1)
        self.sheet_url_text.insert("1.0", self.SHEET_URL)
        self.sheet_url_text.grid(row=0, column=1, padx=10, pady=10)

        self.before_replacement_entry = ttk.Entry(self.root, width=50, font=custom_font, state="readonly")
        self.before_replacement_entry.insert(0, self.BEFORE_REPLACEMENT)
        self.before_replacement_entry.grid(row=1, column=1, padx=10, pady=10)

        self.after_replacement_entry = ttk.Entry(self.root, width=50, font=custom_font, state="readonly")
        self.after_replacement_entry.insert(0, self.AFTER_REPLACEMENT)
        self.after_replacement_entry.grid(row=2, column=1, padx=10, pady=10)

        # Buttons
        self.save_button = ttk.Button(self.root, text="Save Settings", command=self.save_settings)
        self.save_button.grid(row=3, column=0, padx=10, pady=10, sticky="ew")

        self.pause_button = ttk.Button(self.root, text="Pause", command=self.toggle_pause)
        self.pause_button.grid(row=3, column=1, padx=10, pady=10, sticky="ew")

        self.exit_button = ttk.Button(self.root, text="Exit to Reload data", command=self.exit_program)
        self.exit_button.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        # Language dropdown
        self.saved_language = self.config.get("Settings", "language", fallback="vi")

        if not os.path.exists(self.LANGUAGES_FOLDER):
            os.makedirs(self.LANGUAGES_FOLDER)
            self.download_and_process_xlsx_for_languages(self.LINK_EDIT_FILE, self.LANGUAGES_FOLDER)

        self.load_language(self.saved_language)
        language_files = [f for f in os.listdir(self.LANGUAGES_FOLDER) if f.endswith(".ini")]
        language_codes = [os.path.splitext(f)[0] for f in language_files]

        self.language_var = tk.StringVar(self.root)
        self.language_var.set(self.saved_language)
        self.language_menu = ttk.OptionMenu(self.root, self.language_var, self.saved_language, *language_codes, command=self.change_language)
        self.language_menu.grid(row=4, column=2, padx=10, pady=10, sticky="ew")

        # New text field for the "Link Edit File"
        self.link_edit_file_label = ttk.Label(self.root, text="Link Edit File:", font=custom_font)
        self.link_edit_file_label.grid(row=5, column=0, padx=10, pady=10, sticky="w")

        self.link_edit_file_text = tk.Text(self.root, width=50, height=1, font=custom_font, relief="solid", bd=1, state="disabled")
        self.link_edit_file_text.grid(row=5, column=1, padx=10, pady=10)

        self.open_sheet_button = ttk.Button(self.root, text="Open Google Sheet", command=self.open_google_sheet)
        self.open_sheet_button.grid(row=5, column=2, padx=10, pady=10, sticky="ew")

        # Update label to show update status
        self.update_label = ttk.Label(self.root, text="", font=custom_font)
        self.update_label.grid(row=6, column=0, columnspan=2, padx=10, pady=10, sticky="w")

        # Download button (initially hidden)
        self.download_button = ttk.Button(self.root, text="Download Update")
        self.download_button.grid(row=6, column=2, padx=10, pady=10, sticky="ew")

        self.change_language(self.saved_language)
        self.update_gui_language()

        self.sheet_url_text.bind("<FocusOut>", lambda event: self.save_settings())

        # --- Status Bar ---
        self.status_bar = ttk.Label(self.root, text="Ready", relief=tk.SUNKEN, anchor=tk.W, font=custom_font)
        self.status_bar.grid(row=7, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

        # Redirect logging to the status bar
        self.status_bar_handler = self.StatusBarHandler(self.status_bar)
        self.status_bar_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
        logging.getLogger().addHandler(self.status_bar_handler)

        # --- Initialization ---
        self.check_for_new_update(self.update_label, self.download_button)

        # Define constants manually
        self.CF_UNICODETEXT = 13
        self.CF_TEXT = 1

        # Start program logic
        self.replacement_data = self.load_settings_and_data()
        if self.replacement_data:
            self.start_keyboard_listener(self.replacement_data)
            self.clipboard_thread = threading.Thread(
                target=self.monitor_clipboard, args=(self.replacement_data,), daemon=True
            )
            self.clipboard_thread.start()
            logging.info("Clipboard monitoring started.")
            try:
                self.download_and_process_xlsx_for_languages(
                    self.LINK_LANGUAGE_FILE, self.LANGUAGES_FOLDER
                )
                self.load_language(self.saved_language)
                self.update_gui_language()
            except Exception as e:
                logging.error(f"Error downloading or processing language XLSX file: {e}")
             # Update GUI fields after loading data
            self.update_replacement_fields()
            self.update_link_edit_file_field()
        else:
            logging.error("Failed to start program. Exiting program.")

    def get_base_path(self):
        if getattr(sys, 'frozen', False):
            if hasattr(sys, '_MEIPASS'):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))
        return base_path

    # --- src/processXlsx.py content ---
    def value_to_string(self, value):
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            else:
                return str(value)
        elif value is None:
            return ''
        else:
            return str(value)

    def parse_xlsx_for_replacements(self, workbook):
        sheet = workbook.active
        replacement_data = {}

        for row in sheet.iter_rows(min_row=2):
            if len(row) >= 2:
                word_cell = row[0]
                replacement_cell = row[1]
                word = self.value_to_string(word_cell.value).strip()
                replacement = self.value_to_string(replacement_cell.value)

                if word == "BEFORE_REPLACEMENT":
                    self.BEFORE_REPLACEMENT = replacement
                    logging.info(f"Set BEFORE_REPLACEMENT to: '{self.BEFORE_REPLACEMENT}'")
                    continue
                elif word == "AFTER_REPLACEMENT":
                    self.AFTER_REPLACEMENT = replacement
                    logging.info(f"Set AFTER_REPLACEMENT to: '{self.AFTER_REPLACEMENT}'")
                    continue
                elif word == "LINK_EDIT_FILE":
                    self.LINK_EDIT_FILE = replacement
                    logging.info(f"Set LINK_EDIT_FILE to: '{self.LINK_EDIT_FILE}'")
                    continue

                if word and replacement is not None:
                    replacement = replacement.replace('\\n', '\n')
                    replacement_data[word] = replacement
        return replacement_data

    def load_xlsx_from_url(self, xlsx_url):
        with requests.Session() as s:
            download = s.get(xlsx_url, allow_redirects=True, stream=True, timeout=5)
            download.raise_for_status()

            temp_file = "temp_replacement_data.xlsx"
            with open(temp_file, 'wb') as f:
                f.write(download.content)

            try:
                workbook = openpyxl.load_workbook(temp_file)
                replacement_data = self.parse_xlsx_for_replacements(workbook)

                os.remove(temp_file)
                logging.info(f"Set LINK_EDIT_FILE to: '{self.LINK_EDIT_FILE}'")
                return replacement_data
            except Exception as e:
                logging.error(f"Error loading XLSX file: {e}")
                return {}

    def load_xlsx_from_file(self, file_path):
        try:
            workbook = openpyxl.load_workbook(file_path)
            return self.parse_xlsx_for_replacements(workbook)
        except Exception as e:
            logging.error(f"Error loading replacement data from local backup: {e}")
            return {}

    def save_xlsx_to_file(self, replacement_data, file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Sheet1", 0)
        sheet.append(["Word", "Replacement"])

        sheet.append(["BEFORE_REPLACEMENT", self.BEFORE_REPLACEMENT])
        after_replacement_to_save = self.AFTER_REPLACEMENT.replace(' ', '\u00A0')
        sheet.append(["AFTER_REPLACEMENT", after_replacement_to_save])
        sheet.append(["LINK_EDIT_FILE", self.LINK_EDIT_FILE])

        for word, replacement in replacement_data.items():
            sheet.append([word, replacement])

        workbook.save(file_path)

    # --- src/languages.py content ---
    def load_language(self, language_code):
        language_file = os.path.join(self.LANGUAGES_FOLDER, f"{language_code}.ini")

        if os.path.exists(language_file):
            self.language_config.read(language_file, encoding='utf-8')
            logging.info(f"Language set to {language_code}")
        else:
            logging.error(f"Language file for {language_code} not found. Using default language.")

    def download_and_process_xlsx_for_languages(self, xlsx_url, languages_folder):
        try:
            response = requests.get(xlsx_url)
            response.raise_for_status()

            temp_file = os.path.join(languages_folder, "temp_languages.xlsx")
            with open(temp_file, 'wb') as f:
                f.write(response.content)

            workbook = openpyxl.load_workbook(temp_file)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                ini_content = []
                for row in sheet.iter_rows(values_only=True):
                    if row[0]:
                        ini_content.append(row[0])

                ini_file_path = os.path.join(languages_folder, f"{sheet_name}.ini")
                with open(ini_file_path, 'w', encoding='utf-8') as ini_file:
                    ini_file.write('\n'.join(ini_content))

            os.remove(temp_file)
            logging.info("Languages downloaded and processed successfully.")
        except Exception as e:
            logging.error(f"Error downloading or processing XLSX file: {e}")

    # --- src/keyboardEvent.py content ---
    def replace_word(self, word, replacement):
        logging.info(f"Replacing '{word}' with '{replacement}'")
        keyboard.write("\b" * (len(word) + 1))

        replacement_parts = replacement.split('\n')
        keyboard.write(self.BEFORE_REPLACEMENT)
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(replacement, win32clipboard.CF_UNICODETEXT)
        win32clipboard.CloseClipboard()
        keyboard.press_and_release('ctrl+v')

        keyboard.write(self.AFTER_REPLACEMENT)

    def on_mouse_move(self, x, y):
        if abs(x - self.last_mouse_position[0]) > 10 or abs(y - self.last_mouse_position[1]) > 10:
            self.mouse_moved_significantly = True
            self.last_mouse_position = (x, y)

    def on_key_event(self, replacement_data):
        buffer = []
        replacement_words = set(replacement_data.keys())

        def handle_key(event):
            nonlocal buffer

            if self.is_paused:
                return

            if event.event_type == keyboard.KEY_DOWN:
                if event.name in ['space', 'enter']:
                    current_selection = pyperclip.paste().strip()
                    if current_selection:
                        if current_selection in replacement_words:
                            self.replace_word(current_selection, replacement_data[current_selection])
                            buffer = []
                            pyperclip.copy('')
                        else:
                            word = "".join(buffer)
                            if word in replacement_words:
                                self.replace_word(word, replacement_data[word])
                            buffer = []
                    else:
                        if self.mouse_moved_significantly:
                            buffer = []
                            self.mouse_moved_significantly = False
                        word = "".join(buffer)
                        if word in replacement_words:
                            self.replace_word(word, replacement_data[word])
                        buffer = []
                elif event.name == 'backspace':
                    if buffer:
                        buffer.pop()
                elif event.name.isalnum() and event.name not in ['ctrl', 'enter', 'capslock', 'shift']:
                    buffer.append(event.name)
        return handle_key

    def start_keyboard_listener(self, replacement_data):
        if self.keyboard_thread_running:
            self.stop_keyboard_hook()
        self.stop_event.clear()
        keyboard_thread = threading.Thread(target=self.start_keyboard_hook, args=(replacement_data,))
        keyboard_thread.daemon = True
        keyboard_thread.start()
        self.keyboard_thread_running = True
        logging.info("Listening for keyboard input...")

    def start_keyboard_hook(self, replacement_data):
        keyboard.hook(self.on_key_event(replacement_data))

        while not self.stop_event.is_set():
            time.sleep(self.TIME_INTERVAL_KEYBOARD_CHECK)

        keyboard.unhook_all()
        logging.info("Keyboard hook stopped.")

    def stop_keyboard_hook(self):
        self.stop_event.set()

    # --- main.py content ---
    def start_program(self):
        pass  # Logic moved to __init__

    def is_clipboard_text(self):
        try:
            win32clipboard.OpenClipboard()
            if win32clipboard.IsClipboardFormatAvailable(self.CF_UNICODETEXT):
                data = win32clipboard.GetClipboardData(self.CF_UNICODETEXT)
                return True, data
            elif win32clipboard.IsClipboardFormatAvailable(self.CF_TEXT):
                data = win32clipboard.GetClipboardData(self.CF_TEXT)
                return True, data
            else:
                return False, None
        except Exception as e:
            print(f"An error occurred: {e}")
            return False, None
        finally:
            win32clipboard.CloseClipboard()

    def monitor_clipboard(self, replacement_data):
        while True:
            try:
                is_text, content = self.is_clipboard_text()
                if is_text is False:
                    time.sleep(self.TIME_INTERVAL_CLIPBOARD_CHECK)
                    continue
                keys_of_replacement_data = list(replacement_data.keys())
                if content is not None:
                    content = content.strip()
                if content in keys_of_replacement_data:
                    print(is_text, content)
                    new_content = replacement_data[content]
                    logging.info(f"Clipboard content: {new_content}")
                    win32clipboard.OpenClipboard()
                    win32clipboard.EmptyClipboard()
                    win32clipboard.SetClipboardText(
                        new_content, win32clipboard.CF_UNICODETEXT
                    )
                    win32clipboard.CloseClipboard()
                    keyboard.press_and_release("ctrl+v")
            except TypeError:
                pass
            except win32clipboard.error as e:
                logging.error(f"Clipboard error: {e}")
            except Exception as e:
                logging.exception("An unexpected error occurred during clipboard monitoring:")

            time.sleep(self.TIME_INTERVAL_CLIPBOARD_CHECK)

    # --- tkGUI.py methods ---
    def save_settings(self):
        current_sheet_url = self.sheet_url_text.get("1.0", "end-1c").strip()

        if current_sheet_url != self.previous_sheet_url:
            self.previous_sheet_url = current_sheet_url

            self.config["Settings"] = {
                "sheet_url": current_sheet_url,
                "language": self.saved_language,
            }
            with open("settings.ini", "w") as configfile:
                self.config.write(configfile)
            logging.info("Settings saved successfully!")

    def load_replacement_data(self, xlsx_url):
        try:
            replacement_data = self.load_xlsx_from_url(xlsx_url)
            if replacement_data:
                self.save_xlsx_to_file(replacement_data, self.BACKUP_XLSX_PATH)
                logging.info(
                    "Replacement data loaded from the internet and saved to local backup."
                )
            return replacement_data

        except requests.exceptions.RequestException as e:
            logging.error(f"Error downloading data: {e}")
            logging.info("Attempting to load replacement data from local backup...")

            if os.path.exists(self.BACKUP_XLSX_PATH):
                return self.load_xlsx_from_file(self.BACKUP_XLSX_PATH)
            else:
                logging.error("No local backup file found. Exiting program.")
                return {}

    def load_settings_and_data(self):
        self.config.read("settings.ini")
        self.SHEET_URL = self.config.get("Settings", "sheet_url", fallback=self.DEFAULT_SHEET_URL)

        replacement_data = self.load_replacement_data(self.SHEET_URL)
        logging.info(f"Loaded replacement data: {self.LINK_EDIT_FILE} items.")
        if replacement_data:
            self.current_replacement_data = replacement_data
            logging.info("Replacement data reloaded successfully.")

            self.update_replacement_fields()
            self.update_link_edit_file_field()
            logging.info(f"Settings and data loaded successfully.{self.LINK_EDIT_FILE}")
            return replacement_data
        else:
            logging.error("Failed to reload replacement data.")
            return None

    def change_language(self, language_code):
        self.saved_language = language_code
        self.load_language(language_code)
        self.update_gui_language()

        self.config["Settings"] = {
            "language": language_code,
            "sheet_url": self.sheet_url_text.get("1.0", "end-1c"),
        }
        with open("settings.ini", "w") as configfile:
            self.config.write(configfile)
        logging.info(f"Language changed to {language_code} and saved to settings file.")

    def update_gui_language(self):
        try:
            self.sheet_url_label.config(text=self.language_config["Labels"]["sheet_url"])
            self.before_replacement_label.config(text=self.language_config["Labels"]["before_replacement"])
            self.after_replacement_label.config(text=self.language_config["Labels"]["after_replacement"])
            self.link_edit_file_label.config(text=self.language_config["Labels"]["link_edit_file"])

            self.save_button.config(text=self.language_config["Buttons"]["save_settings"])
            self.pause_button.config(text=self.language_config["Buttons"]["pause"] if not self.is_paused else self.language_config["Buttons"]["resume"])
            self.exit_button.config(text=self.language_config["Buttons"]["exit_to_reload_data"])
            self.open_sheet_button.config(text=self.language_config["Buttons"]["open_google_sheet"])
            self.download_button.config(text=self.language_config["Buttons"]["download_update"])

        except KeyError as e:
            logging.error(f"Missing translation for key: {e}")

    def update_replacement_fields(self):
        self.before_replacement_entry.config(state="normal")
        self.before_replacement_entry.delete(0, tk.END)
        self.before_replacement_entry.insert(0, self.BEFORE_REPLACEMENT)
        self.before_replacement_entry.config(state="readonly")

        self.after_replacement_entry.config(state="normal")
        self.after_replacement_entry.delete(0, tk.END)
        self.after_replacement_entry.insert(0, self.AFTER_REPLACEMENT)
        self.after_replacement_entry.config(state="readonly")

    def update_link_edit_file_field(self):
        self.link_edit_file_text.config(state="normal")
        self.link_edit_file_text.delete("1.0", tk.END)
        self.link_edit_file_text.insert("1.0", self.LINK_EDIT_FILE)
        self.link_edit_file_text.config(state="disabled")

    def toggle_pause(self):
        self.is_paused = not self.is_paused
        if self.is_paused:
            self.pause_button.config(text=self.language_config["Buttons"]["resume"])
            logging.info("Program paused.")
        else:
            self.pause_button.config(text=self.language_config["Buttons"]["pause"])
            self.style.map("TButton", background=[("active", "darkgreen")])
            logging.info("Program resumed.")

    def exit_program(self):
        self.root.quit()
        self.root.destroy()
        sys.exit()

    def open_google_sheet(self):
        url = self.link_edit_file_text.get("1.0", "end-1c").strip()
        if url:
            webbrowser.open(url)
        else:
            messagebox.showwarning("No URL", "The 'Link Edit File' URL is not set.")

    # --- src/updater.py content ---
    def get_current_version(self):
        return self.VERSION

    def check_for_new_update(self, update_label, download_button):
        current_version = self.get_current_version()
        update_label.config(text=f"Current version: {current_version}")

        for update_url in self.UPDATE_URLS:
            try:
                version_url = f"{update_url}/version.txt"
                response = requests.get(version_url, timeout=self.HTTP_TIMEOUT)
                response.raise_for_status()
                latest_version = response.text.strip()

                if version.parse(latest_version) > version.parse(current_version):
                    update_label.config(text=f"New version available: {latest_version}")
                    download_button.config(
                        command=lambda: self.open_download_page(update_url),
                        state=tk.NORMAL,
                    )
                    download_button.config(state=tk.NORMAL)
                    return
                else:
                    update_label.config(text="You are on lastest version.")
                    download_button.grid_forget()
                    return

            except requests.exceptions.RequestException:
                pass
            except Exception:
                pass

    def open_download_page(self, update_url):
        webbrowser.open_new_tab(update_url)

    class StatusBarHandler(logging.Handler):
        def __init__(self, status_bar):
            logging.Handler.__init__(self)
            self.status_bar = status_bar

        def emit(self, record):
            log_entry = self.format(record)
            self.status_bar.after(
                0, self.status_bar.config, {"text": log_entry}
            )

if __name__ == "__main__":
    app = TextReplacerApp()
    atexit.register(app.stop_keyboard_hook)
    mouse_listener = mouse.Listener(on_move=app.on_mouse_move)
    mouse_listener.start()
    app.root.mainloop()